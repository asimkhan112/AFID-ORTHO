"""Excel (.xlsx) generation service using openpyxl.

Kept separate from the route/DB layer: the route builds a list of plain row
dicts from PostgreSQL and hands them here, so this module has no database or
FastAPI dependencies and is easy to test in isolation.
"""
from io import BytesIO
from datetime import date
from typing import List, Dict
import logging

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column order is the export contract expected by the client.
COLUMNS = [
    ("queue_number", "Queue Number", 14),
    ("patient_id", "Patient ID", 16),
    ("patient_name", "Patient Name", 26),
    ("age", "Age", 8),
    ("gender", "Gender", 10),
    ("visit_date", "Visit Date", 14),
    ("visit_time", "Visit Time", 12),
    ("diagnosis", "Diagnosis", 26),
    ("procedure", "Procedure/Treatment", 22),
    ("status", "Status", 14),
    ("doctor_name", "Doctor Name", 24),
]

_HEADER_FILL = PatternFill("solid", fgColor="123B64")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, color="123B64", size=13)
_THIN = Side(style="thin", color="D9E2EC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_daily_queue_workbook(rows: List[Dict], for_date: date, doctor_name: str) -> bytes:
    """Build the daily-queue workbook and return its raw .xlsx bytes.

    `rows` is a list of dicts keyed by the first element of each COLUMNS entry.
    The workbook is written to an in-memory buffer (no temp files), so the
    caller can stream it straight back in the HTTP response.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Queue"

    header_labels = [label for _, label, _ in COLUMNS]
    n_cols = len(COLUMNS)

    # ── Title block ──────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=f"Daily Patient Queue — {doctor_name}")
    title_cell.font = _TITLE_FONT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub_cell = ws.cell(row=2, column=1, value=f"Date: {for_date.isoformat()}   ·   Records: {len(rows)}")
    sub_cell.font = Font(color="667085", size=10)

    header_row_idx = 4

    # ── Header row ───────────────────────────────────────────────────────────
    for col_idx, label in enumerate(header_labels, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER

    # ── Data rows ────────────────────────────────────────────────────────────
    keys = [key for key, _, _ in COLUMNS]
    for r_offset, row in enumerate(rows, start=header_row_idx + 1):
        for col_idx, key in enumerate(keys, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=r_offset, column=col_idx, value=value if value not in (None, "") else "N/A")
            cell.alignment = Alignment(horizontal="center" if key in {"queue_number", "age", "gender", "visit_time", "status"} else "left", vertical="center")
            cell.border = _BORDER

    # ── Column widths + freeze header ───────────────────────────────────────
    for col_idx, (_, _, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    logger.info("Generated daily-queue workbook: %d rows for %s (%s)", len(rows), doctor_name, for_date.isoformat())
    return buffer.getvalue()
